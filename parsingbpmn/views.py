import csv
from datetime import datetime

from django.core.files.storage import FileSystemStorage
from django.core.serializers import python
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

from .forms import ProcessForm, SystemForm
from .models import Process, Asset, System, Asset_has_attribute, Attribute, Asset_type, Attribute_value, \
    Threat_has_attribute, Threat_has_control, ThreatAgentRiskScores, TACategoryAttribute, ThreatAgentCategory, \
    System_ThreatAgent, TAReplies_Question, TAReplyCategory, Reply, ThreatAgentQuestion
from .bpmn_python_master.bpmn_python import bpmn_diagram_rep as diagram

# Create your views here.

def system_management(request):
    if request.method == 'POST':
        form = SystemForm(request.POST)
        if form.is_valid():
            form.save()
            last_system = System.objects.latest('id')
            return redirect('bpmn_process_management', last_system.pk)
    else:
        form = SystemForm()
    systems = System.objects.all()
    return render(request,'system_management.html',{
        'form':form,'systems':systems
    })

def bpmn_process_management(request,pk):
    if request.method == 'POST':
        form = ProcessForm(request.POST, request.FILES)
        if form.is_valid():
            saved_form = form.save(commit=False)
            saved_form.system_id = pk
            saved_form.save()
            last_process = Process.objects.latest('id')
            bpmn_graph = diagram.BpmnDiagramGraph()
            pk = last_process.pk
            bpmn_graph.load_diagram_from_xml_file(Process.objects.get(pk=pk).xml)
            lista = bpmn_graph.get_nodes()
            #print(lista)
            annotations=[]
            associations=[]

            for tuple in lista:
                for dizionario in tuple:
                    if type(dizionario) is dict:
                        try:
                            if dizionario['type'].endswith("textAnnotation"):
                                annotations.append(dizionario)
                        except KeyError:
                            print()
                        try:
                            if dizionario['type'].endswith("association"):
                                associations.append(dizionario)
                        except KeyError:
                            print()

            e=""
            for tuple in lista:
                for dizionario in tuple:
                    if type(dizionario) is dict:
                        if dizionario['type'].endswith("Task"):
                            attribute_value = []
                            id_task = dizionario['id']
                            x = dizionario["x"]
                            y = dizionario["y"]
                            width = dizionario["width"]
                            height = dizionario["height"]
                            position = x + ":" + y + ":" + width + ":" + height
                            if dizionario['type'].startswith("send"):
                                asset_type = Asset_type.objects.get(name="Send task")
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref=assoc['association'][2]
                                        for textAnn in annotations:
                                            if(target_ref==textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                    e = e.lower()
                                if e=="pec_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="PEC communication"))
                                elif e=="mail_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Mail communication"))
                                elif e=="post_office_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Post office communication"))
                            elif dizionario['type'].startswith("receive"):
                                asset_type = Asset_type.objects.get(name="Receive task")
                                id_task = dizionario['id']
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                    e = e.lower()
                                if e == "pec_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="PEC communication"))
                                elif e == "mail_communication":
                                    attribute_value.append(Attribute_value.objects.get(value="Mail communication"))
                                elif e == "post_office_communication":
                                    attribute_value.append(
                                        Attribute_value.objects.get(value="Post office communication"))
                            elif dizionario['type'].startswith("user"):
                                asset_type = Asset_type.objects.get(name="User task")
                                id_task = dizionario['id']
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                    e = e.lower()
                                if e=="online":
                                    attribute_value.append(Attribute_value.objects.get(value="Online"))
                                elif e=="offline":
                                    attribute_value.append(Attribute_value.objects.get(value="Offline"))
                            elif dizionario['type'].startswith("manual"):
                                asset_type = Asset_type.objects.get(name="Manual task")
                                attribute_value.append(Attribute_value.objects.get(value="Manual task"))
                            elif dizionario['type'].startswith("service"):
                                asset_type = Asset_type.objects.get(name="Service task")
                                id_task = dizionario['id']
                                e=""
                                for assoc in associations:
                                    if(id_task==assoc['association'][1]):
                                        target_ref = assoc['association'][2]
                                        for textAnn in annotations:
                                            if (target_ref == textAnn['id']):
                                                e= (textAnn['textAnnotation'][0][1])
                                    e = e.replace(" ","_")
                                e = e.lower()
                                if e=="statefull":
                                    attribute_value.append(Attribute_value.objects.get(value="Statefull"))
                                elif e=="stateless":
                                    attribute_value.append(Attribute_value.objects.get(value="Stateless"))
                            elif dizionario['type'].startswith("script"):
                                asset_type = Asset_type.objects.get(name="Script task")
                                attribute_value.append(Attribute_value.objects.get(value="Script task"))
                            elif dizionario['type'].startswith("business"):
                                asset_type = Asset_type.objects.get(name="Business rule task")
                                attribute_value.append(Attribute_value.objects.get(value="Business rule task"))
                            asset = Asset(name=dizionario['node_name'],bpmn_id=id_task,position=position, process=Process.objects.get(pk=pk),asset_type=asset_type)
                            asset.save()
                            attribute = []
                            for value in attribute_value:
                                attribute.append(Attribute.objects.get(asset_type=asset_type,attribute_value=value))
                            for a in attribute:
                                asset_has_attribute = Asset_has_attribute(asset=asset,attribute=a)
                                asset_has_attribute.save()
                        elif dizionario['type'].endswith("task"):
                            asset = Asset(name=dizionario['node_name'], process=Process.objects.get(pk=pk))
                            asset.save()



            return redirect('process_view_task_type', pk)
    else:
        form = ProcessForm()
    processes = Process.objects.filter(system=System.objects.get(pk=pk))
    check_box = []
    for process in processes:
        assets = Asset.objects.filter(process=process)
        check_attribute = False
        for asset in assets:
            if not Asset_has_attribute.objects.filter(asset=asset):
                check_attribute = True
        check_box.append(check_attribute)
    processes_info = zip(processes,check_box)
    return render(request,'bpmn_process_management.html',{
        'form':form,'processes_info':processes_info,'pk':pk, 'processes':processes
    })

def delete_system(request,pk):
    if request.method == 'POST':
        system = System.objects.get(pk=pk)
        system.delete()
    return redirect('system_management')

def delete_process(request,pk):
    if request.method == 'POST':
        process = Process.objects.get(pk=pk)
        system_id = process.system.pk
        process.delete()
    return redirect('bpmn_process_management',system_id)

def process_view_task_type(request,pk):
    task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
    check_attribute = False
    for task in task_list:
        if task.asset_type == None:
            check_attribute = True
    if check_attribute == True:
        asset_type = Asset_type.objects.all()
        system = Process.objects.get(pk=pk).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_task_type.html', {
            'task_list':task_list,'asset_type':asset_type,'pk':pk,'processes':processes
        })
    else:
        return redirect('process_view_attribute', pk)

def task_type_enrichment(request,pk):
    if request.method == "POST":
        assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
        task_enrichment = []
        types = []
        for asset in assets_for_process:
            task_enrichment.append(request.POST.get(str(asset.pk)))
        for type in task_enrichment:
            if type != None:
                type = int(type)
                types.append(Asset_type.objects.get(pk=type))
            else:
                types.append(None)
        for asset,type in zip(assets_for_process,types):
            if type != None:
                x = Asset.objects.get(pk=asset.pk)
                x.asset_type = type
                x.save()
        return redirect('process_view_attribute',pk)
    else:
        return redirect('task_type_enrichment',pk)


def process_view_attribute(request,pk):
    task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
    check_attribute = False
    for task in task_list:
        if not Asset_has_attribute.objects.filter(asset=task):
            check_attribute = True
    if check_attribute==True:
        task_attributes = []
        list_attributes = []
        for task in task_list:
            task_attributes.append(Asset_has_attribute.objects.filter(asset=task))
        for attributes in task_attributes:
            if not attributes:
                list_attributes.append("empty")
            else:
                sub_list = []
                for element in attributes:
                    sub_list.append(element.attribute.attribute_value.value)
                list_attributes.append(sub_list)
        send = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Send task"))
        receive = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Receive task"))
        user = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="User task"))
        manual = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Manual task"))
        service = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Service task"))
        script = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Script task"))
        business = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Business rule task"))
        task_info = zip(task_list,list_attributes)
        system = Process.objects.get(pk=pk).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_attribute.html', {
                'task_info':task_info,'send':send,'receive':receive,'user':user,'manual':manual,'service':service,
                'script':script,'business':business,'pk':pk,'processes':processes})
    else:
        return redirect('threats_and_controls',pk)

def process_enrichment(request,pk):
    if request.method == "POST":
        task_list = Asset.objects.filter(process=Process.objects.get(pk=pk))
        pathfile=Process.objects.filter(id=pk)[0].xml

        check_attribute = False
        for task in task_list:
            if not Asset_has_attribute.objects.filter(asset=task):
                check_attribute = True
        if check_attribute == True:
            assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
            attributes_enrichment = []
            attributes = []
            for asset in assets_for_process:
                attributes_enrichment.append(request.POST.get(str(asset.pk)))
            for attribute_enrichment in attributes_enrichment:
                if attribute_enrichment != None:
                    attribute_enrichment = int(attribute_enrichment)
                    attributes.append(Attribute.objects.get(pk=attribute_enrichment))
                else:
                    attributes.append(None)

            for asset,attribute in zip(assets_for_process,attributes):
                if attribute != None:
                    asset_has_attribute = Asset_has_attribute(asset=asset,attribute=attribute)

                    writeTextAnnotation_bpmn(pathfile,asset.position,asset.bpmn_id,attribute.attribute_value)


                    asset_has_attribute.save()

            return redirect('threats_and_controls',pk)
        else:
            assets_for_process = Asset.objects.filter(process=Process.objects.get(pk=pk))
            attributes_enrichment = []
            attributes = []
            for asset in assets_for_process:
                attributes_enrichment.append(request.POST.get(str(asset.pk)))
            for attribute_enrichment in attributes_enrichment:
                if attribute_enrichment != None:
                    attribute_enrichment = int(attribute_enrichment)
                    attributes.append(Attribute.objects.get(pk=attribute_enrichment))
                else:
                    attributes.append(None)

            for asset, attribute in zip(assets_for_process, attributes):
                if attribute != None:
                    Asset_has_attribute.objects.filter(asset=asset).update(attribute=attribute)
            return redirect('threats_and_controls', pk)
    else:
        return redirect('process_enrichment',pk)

import random
import string

def get_random_string(length):
    letters = string.ascii_letters
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str

def writeTextAnnotation_bpmn(pathfile,position,taskId,attribute_value):
    textAnnotationId="TextAnnotation_"+get_random_string(7)
    textAnnotation="\
    <bpmn:textAnnotation id=\""+str(textAnnotationId)+"\">\n \
    <bpmn:text>"+str(attribute_value)+"</bpmn:text>\n \
    </bpmn:textAnnotation>\n"
    associationId="Association_"+get_random_string(7)
    association="<bpmn:association id=\""+str(associationId)+"\" sourceRef=\""+taskId+"\" targetRef=\""+textAnnotationId+"\" />"
    positionValues=position.split(":")
    x=positionValues[0]
    y=positionValues[1]
    width=positionValues[2]
    height=positionValues[3]

    stringToWrite=str(textAnnotation)+" "+str(association)
    shapetextAnn="<bpmndi:BPMNShape id=\""+textAnnotationId+"_di\" bpmnElement="+textAnnotationId+">\n\
        <dc:Bounds x=\""+str(int(x)+20)+"\" y=\""+str(int(y)+20)+"\" width=\""+str(width)+"\" height=\""+height+"\" />\n\
      </bpmndi:BPMNShape>\n"

    shapeAssoc="<bpmndi:BPMNShape id=\""+associationId+"_di\" bpmnElement="+associationId+">\n\
        <dc:Bounds x=\""+x+"\" y=\""+y+"\"/>\n\
        <dc:Bounds x=\""+str(int(x)+20)+"\" y=\""+str(int(y)+20)+"\"/>\n\
      </bpmndi:BPMNShape>\n"

    f = open(str(pathfile), "r+")
    stringFile=f.read()

    from xml.dom.minidom import parse, parseString

    datasource = open(str(pathfile)) #convert to minidom object
    minidomObject = parse(datasource)
    process = minidomObject.getElementsByTagName('bpmn:process')
    for e in process:
        if(taskId in e.toxml()):

            print("per il task "+str(taskId)+" il process è ")
            print(e.toxml())

    #task[0].firstChild.nodeValue

    #Process_1j43nxw


    #print(minidomObject.toxml()) #convert to xml string





def edit_process(request,pk):
    if request.method == "POST":
        assets = Asset.objects.filter(process=Process.objects.get(pk=pk))
        assets_type = []
        list_attributes = []
        for asset in assets:
            assets_type.append(asset.asset_type)
            list_attributes.append("empty")

        task_info = zip(assets,list_attributes)
        send = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Send task"))
        receive = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Receive task"))
        user = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="User task"))
        manual = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Manual task"))
        service = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Service task"))
        script = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Script task"))
        business = Attribute.objects.filter(asset_type=Asset_type.objects.get(name="Business rule task"))
        system = Process.objects.get(pk=pk).system
        processes = Process.objects.filter(system=system)
        return render(request, 'process_view_attribute.html', {
            'task_info': task_info, 'send': send, 'receive': receive, 'user': user, 'manual': manual,
            'service': service,'script': script, 'business': business, 'pk': pk, 'processes': processes})

def threats_and_controls(request,pk):
    process = Process.objects.get(pk=pk)
    assets = Asset.objects.filter(process=process)
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            sublist_controls.append(Threat_has_control.objects.filter(threat=threat))
        controls.append(sublist_controls)

    clear_list_threats = []
    for threat_list in threats:
        for threat in threat_list:
            if threat.threat not in clear_list_threats:
                clear_list_threats.append(threat.threat)

    clear_list_controls = []
    for control_of_asset in controls:
        for control_of_threat in control_of_asset:
            for control in control_of_threat:
                if control.control not in clear_list_controls:
                    clear_list_controls.append(control.control)

    system = Process.objects.get(pk=pk).system
    processes = Process.objects.filter(system=system)
    return render(request, 'threats_and_controls.html', {
        'process_name':process.name,'clear_list_threats': clear_list_threats,'clear_list_controls':clear_list_controls,'pk':pk,'processes':processes
    })

def threat_modeling(request,pk):
    assets = Asset.objects.filter(process=Process.objects.get(pk=pk))
    attributes = []
    threats = []
    controls = []
    for asset in assets:
        attributes.append(Asset_has_attribute.objects.filter(asset=asset))
    for list_attribute in attributes:
        for attribute in list_attribute:
            attribute = attribute.attribute
            threats.append(Threat_has_attribute.objects.filter(attribute=attribute))
    for threats_of_asset in threats:
        sublist_controls = []
        for threat in threats_of_asset:
            threat = threat.threat
            sublist_controls.append(Threat_has_control.objects.filter(threat=threat))
        controls.append(sublist_controls)

    controls_per_asset = []
    for asset in threats:
        list_controls = []
        for threat in asset:
            threat = threat.threat
            controls_per_threat = Threat_has_control.objects.filter(threat=threat)
            for control in controls_per_threat:
                control= control.control
                if control not in list_controls:
                    list_controls.append(control)
        controls_per_asset.append(list_controls)

    threat_model_info = zip(assets, attributes, threats, controls,controls_per_asset)
    system = Process.objects.get(pk=pk).system
    processes = Process.objects.filter(system=system)
    return render(request, 'threat_modeling.html',{
        'threat_model_info':threat_model_info,'pk':pk,'processes':processes
    })

def export_threat_modeling(request,pk):
    if request.method == "POST":

        #help: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
            name=Process.objects.get(pk=pk).name.replace(" ","_")
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Threat_modeling_REPORT'
        columns = ['Asset name', 'Asset type', 'Asset attributes', 'Threats','Policy per asset']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman",size=12,bold=True,color='FF0000')
            cell.border = Border(left=Side(border_style="thin",color='FF000000'),
                                 right=Side(border_style="thin",color='FF000000'),
                                 top=Side(border_style="thin",color='FF000000'),
                                 bottom=Side(border_style="thin",color='FF000000'),)

        assets = Asset.objects.filter(process=Process.objects.get(pk=pk))
        attributes = []
        threats = []
        for asset in assets:
            attributes.append(Asset_has_attribute.objects.filter(asset=asset))
        for list_attribute in attributes:
            for attribute in list_attribute:
                attribute = attribute.attribute
                threats.append(Threat_has_attribute.objects.filter(attribute=attribute))

        attributes_list = []
        for attribute in attributes:
            attr_sublist = []
            for element in attribute:
                attr_sublist.append(element.attribute.attribute_value.value)
            attributes_list.append(attr_sublist)

        threats_list = []
        for threat in threats:
            threat_sublist = []
            for element in threat:
                threat_sublist.append(element.threat.name)
            threats_list.append(threat_sublist)

        controls_per_asset = []
        for asset in threats:
            list_controls = []
            for threat in asset:
                threat = threat.threat
                controls_per_threat = Threat_has_control.objects.filter(threat=threat)
                for control in controls_per_threat:
                    control = control.control
                    if control not in list_controls:
                        list_controls.append(control)
            controls_per_asset.append(list_controls)

        for asset,attribute,threat,control in zip(assets,attributes_list,threats_list,controls_per_asset):
            row_num += 1

            if not threat:
                threat0 = ''
            else:
                threat0 = str(threat[0])

            # Define the data for each cell in the row
            row = [
                asset.name,
                asset.asset_type.name,
                str(attribute[0]),
                threat0,
                "CIS."+str(control[0].pk)+" - "+str(control[0])
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )

            count_attr = 0
            old_row = row_num
            while count_attr < len(attribute)-1:
                count_attr += 1
                row_num += 1

                row = [
                    '',
                    '',
                    str(attribute[count_attr]),
                    ''
                ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )

            count_threats = 0
            count_controls = 0
            row_num = old_row
            while count_threats < len(threat)-1 or count_controls < len(control)-1:
                row_num += 1

                if count_threats < len(threat)-1 and count_controls < len(control)-1:
                    count_threats += 1
                    count_controls += 1

                    row = [
                        '',
                        '',
                        '',
                        str(threat[count_threats]),
                        "CIS." + str(control[count_controls].pk) + " - " + str(control[count_controls])
                    ]
                elif count_threats < len(threat)-1 and not count_controls < len(control)-1:
                    count_threats += 1

                    row = [
                        '',
                        '',
                        '',
                        str(threat[count_threats]),
                        ''
                    ]
                else:
                    count_controls += 1

                    row = [
                        '',
                        '',
                        '',
                        '',
                        "CIS." + str(control[count_controls].pk) + " - " + str(control[count_controls])
                    ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )
        #Per effettuare il resize delle celle in base a quella più grande
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value

        workbook.save(response)

        return response

def bpmn_viewer(request,pk):
    process = Process.objects.get(pk=pk)
    return render(request,'bpmn_viewer.html',{
        'process':process
    })


def risk_analysis(request, appId):
    app = System.objects.get(appId=appId)
    appName=app.application
    SelectedComponentName=''
    componentsWithThreats=[]
    components=Asset.objects.filter(app=app)

    try:
        if request.POST['dropdown']:
            SelectedComponentName=request.POST['dropdown']
        else:
            SelectedComponentName=components[0].name
    except:
        print()
    componentUnderAnalysis=components[0]
    for component in components:
        if (len(threat_modeling_per_assetFun(component.id)) != 0):
            if(SelectedComponentName==component.name):
                componentsWithThreats.append((component,True))
                componentUnderAnalysis=component
            else:
                componentsWithThreats.append((component,False))

    threats = threat_modeling_per_assetFun(componentUnderAnalysis.id)

    TAscores=ThreatAgentRiskScores.objects.filter(app=app)

    #ricerca ultimo risultato.
    maxtimeTA = TAscores[0].updated_at
    lastScore=TAscores[0]
    for Tascore in TAscores:
        if(Tascore.updated_at>maxtimeTA):
            lastScore=Tascore

    SIRecords = StrideImpactRecord.objects.filter(app=app)

    PreCondition="[n,n,n]"
    PostCondition="[n,n,n]"
    LossOfConfidentiality=0
    LossOfIntegrity=0
    LossOfAvailability=0
    LossOfCPostConditionValue = 0
    LossOfIPostConditionValue = 0
    LossOfAPostConditionValue = 0
    LossOfCPreConditionValue = 0
    LossOfIPreConditionValue = 0
    LossOfAPreConditionValue = 0

    for threat in threats:
        PreCondition=str(threat[0].PreCondition)
        PostCondition=str(threat[0].PostCondition)
        maxFinancial = 0
        maxReputation = 0
        maxnoncompliance = 0
        maxprivacy = 0
        for SIRecord in SIRecords:
            for Threatstride in threat[1]:
                if(SIRecord.stride.category.lower()==Threatstride.lower()):
                    if(maxFinancial < SIRecord.financialdamage):
                        maxFinancial=SIRecord.financialdamage
                    if (maxReputation < SIRecord.reputationdamage):
                        maxReputation = SIRecord.reputationdamage
                    if (maxnoncompliance < SIRecord.noncompliance):
                        maxnoncompliance = SIRecord.noncompliance
                    if (maxprivacy < SIRecord.privacyviolation):
                        maxprivacy = SIRecord.privacyviolation
        threat[0].financial=maxFinancial
        threat[0].reputation=maxReputation
        threat[0].noncompliance=maxnoncompliance
        threat[0].privacy=maxprivacy

        #elimino [ e ]

        try:
            PreCondition.replace("[","")
            PreCondition.replace("]","")
            PostCondition.replace("[","")
            PostCondition.replace("]","")

            #splitto con le ,
            PreCondition=PreCondition.split(",")
            PostCondition=PostCondition.split(",")

            if(PreCondition[0]=='n'):
                LossOfCPreConditionValue=0
            if (PreCondition[0] == 'p'):
                LossOfCPreConditionValue = 1
            if(PreCondition[0]=='f'):
                LossOfCPreConditionValue=2

            if(PostCondition[0]=='n'):
                LossOfCPostConditionValue=0
            if (PostCondition[0] == 'p'):
                LossOfCPostConditionValue = 1
            if(PostCondition[0]=='f'):
                LossOfCPostConditionValue=2

            LossOfConfidentiality=((LossOfCPostConditionValue+LossOfCPreConditionValue)*3)+1

            if (PreCondition[1] == 'n'):
                LossOfIPreConditionValue = 0
            if (PreCondition[1] == 'p'):
                LossOfIPreConditionValue = 1
            if (PreCondition[1] == 'f'):
                LossOfIPreConditionValue = 2

            if (PostCondition[1] == 'n'):
                LossOfIPostConditionValue = 0
            if (PostCondition[1] == 'p'):
                LossOfIPostConditionValue = 1
            if (PostCondition[1] == 'f'):
                LossOfIPostConditionValue = 2

            LossOfIntegrity = ((LossOfIPostConditionValue + LossOfIPreConditionValue) * 3) + 1

            if (PreCondition[2] == 'n'):
                LossOfAPreConditionValue = 0
            if (PreCondition[2] == 'p'):
                LossOfAPreConditionValue = 1
            if (PreCondition[2] == 'f'):
                LossOfAPreConditionValue = 2
            if (PostCondition[2] == 'n'):
                LossOfAPostConditionValue = 0
            if (PostCondition[2] == 'p'):
                LossOfAPostConditionValue = 1
            if (PostCondition[2] == 'f'):
                LossOfAPostConditionValue = 2

            LossOfAvailability = ((LossOfAPostConditionValue + LossOfAPreConditionValue) * 3) + 1

            threat[0].lossofc=LossOfConfidentiality
            threat[0].lossofi=LossOfIntegrity
            threat[0].lossofa=LossOfAvailability

        except:
            print("iNFO MISSING")

    return render(request, 'risk_analysis.html', {"appName": appName,"ComponentName":SelectedComponentName,"threats":threats,
                                                  "components":componentsWithThreats,"ThreatAgentScores":lastScore,"appId": appId})

@csrf_exempt
def threat_agent_wizard(request,appId):
    context={}
    #Generate question and related replies
    questions=ThreatAgentQuestion.objects.all()
    questions_replies=TAReplies_Question.objects.all()
    questions_replies_list=[]
    for question in questions:
        replies = []
        question_replies_dict = {}
        for reply in questions_replies:
            if question==reply.question:
                replies.append(reply.reply.reply)
        question_replies_dict['question']=question.question
        question_replies_dict['replies']=replies
        questions_replies_list.append(question_replies_dict)
    context['questions_replies']=questions_replies_list
    context['appId']=appId
    return render(request, 'threat_agent_wizard.html', context)

@csrf_exempt
def threat_agent_generation(request, appId):
    print(appId)
    context={}
    ThreatAgents = []
    ThreatAgentsPerAsset = []
    #for category in ThreatAgentCategory.objects.all():   #inizializzo la lista finale a tutti i TA
        #ThreatAgents.append(category)
    for reply in request.POST: #per ogni risposta al questionario
        if(reply!='csrfmiddlewaretoken'):
            ReplyObject=Reply.objects.filter(reply=reply).get()
            tareplycategories=TAReplyCategory.objects.filter(reply=ReplyObject)
            TAList=[]
            for replycategory in tareplycategories.all(): #ogni categoria relativa ad una singola risposta
                #print(replycategory.reply.reply + " "+ replycategory.category.category)
                TAList.append(replycategory.category)
                question = TAReplies_Question.objects.filter(reply=ReplyObject)
            ThreatAgentsPerAsset.append((TAList,question))


    numQ3=0
    numQ4=0
    #conto il numero di risposte date per Q3 e Q4
    for ThreatAgentsList,question in ThreatAgentsPerAsset: #per ogni risposta
        questionId=question.get().question.Qid
        if(questionId=="Q3"):
            numQ3+=1
        if(questionId=="Q4"):
            numQ4+=1

    i=0
    j=0
    ThreatAgentsListTemp=[]
    print(ThreatAgentsPerAsset)
    for ThreatAgentsList,question in ThreatAgentsPerAsset: #per ogni risposta
        questionId=question.get().question.Qid
        if(questionId==1):
            ThreatAgents=ThreatAgentsList
        if(questionId==2):
            ThreatAgents=intersection(ThreatAgents,ThreatAgentsList)
        if(questionId==3):
            if(i==0):
                ThreatAgentsListTemp = ThreatAgentsList
            elif(i<numQ3):
                ThreatAgentsList=union(ThreatAgentsList,ThreatAgentsListTemp)
                ThreatAgentsListTemp=ThreatAgentsList
            if(i==numQ3-1):
                ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
            i = i + 1

        if(questionId==4):
            if(j==0):
                ThreatAgentsListTemp=ThreatAgentsList
                j=j+1
            elif(j==1):
                ThreatAgentsListTemp = ThreatAgentsList
                j=j+1
            elif(j<numQ4):
                ThreatAgentsList=union(ThreatAgentsList,ThreatAgentsListTemp)
                ThreatAgentsListTemp=ThreatAgentsList

    ThreatAgents = intersection(ThreatAgents, ThreatAgentsList)
    print(ThreatAgentsList)
    print(ThreatAgents)
    print(ThreatAgents)
    ThreatAgentsWithInfo={}
    for ta in ThreatAgents:
        ThreatAgentsWithInfo[ta]=list(TACategoryAttribute.objects.filter(category=ta))
        System_ThreatAgent.objects.get_or_create(
            app = System.objects.get(appId=appId),
            category=ta
        )


    context={'ThreatAgents':ThreatAgentsWithInfo}
    context['appId']=appId
    return render(request, 'threat_agent_generation.html',context=context)

def intersection(lst1, lst2):
    lst3 = [value for value in lst1 if value in lst2]
    return lst3

def union(lst1, lst2):
    lst3 = list(set(lst1+lst2))
    return lst3

@csrf_exempt
def calculate_threat_agent_risks(request,appId):
    OWASP_Motive_TOT = 0
    OWASP_Size_TOT = 0
    OWASP_Opportunity_TOT = 0
    OWASP_Skill_TOT = 0
    sommapesi = 0

    for category,risk_value in request.POST.items():
        TACategory=ThreatAgentCategory.objects.get(category=category)
        #per ogni categoria ottieni i Attribute relativi e calcola i 4 parametri owasp con le formule nella tesi.
        TACategoryAttributes=TACategoryAttribute.objects.filter(category=TACategory)
        OWASP_Motive=0
        OWASP_Size=0
        OWASP_Opportunity=0
        OWASP_Skill=0
        limits=0
        intent=0
        access=0
        resources=0
        visibility=0
        skills=0

        OWASP_Motives=[]

        #scorro gli attributi di category
        for TACategoryAttributeVar in TACategoryAttributes:
            if(TACategoryAttributeVar.attribute.attribute=='Skills'):
                skills=TACategoryAttributeVar.attribute.score
            if(TACategoryAttributeVar.attribute.attribute=='Resources'):
                resources=TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Visibility'):
                visibility= TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Limits'):
                limits= TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Intent'):
                intent= TACategoryAttributeVar.attribute.score
            if (TACategoryAttributeVar.attribute.attribute == 'Access'):
                access= TACategoryAttributeVar.attribute.score

        if(risk_value=='L'):
            risk_valueNum= 1
        if (risk_value == 'M'):
            risk_valueNum = 2
        if (risk_value == 'H'):
            risk_valueNum = 3



        sommapesi=sommapesi+risk_valueNum
        OWASP_Motive= ((((intent/2)+(limits/4))/2) * 10)
        OWASP_Opportunity= ((((access/2)+(resources/6)+(visibility/4))/3) * 10)
        OWASP_Size= (resources/6) * 10
        OWASP_Skill= (skills/4) * 10

        OWASP_Motive_TOT += (OWASP_Motive * risk_valueNum)
        OWASP_Opportunity_TOT += OWASP_Opportunity * risk_valueNum
        OWASP_Size_TOT += OWASP_Size * risk_valueNum
        OWASP_Skill_TOT += OWASP_Skill * risk_valueNum

    OWASP_Skill_TOT= int(round(OWASP_Skill_TOT/sommapesi))
    OWASP_Motive_TOT= int(round(OWASP_Motive_TOT/sommapesi))
    OWASP_Size_TOT= int(round(OWASP_Size_TOT/sommapesi))
    OWASP_Opportunity_TOT= int(round(OWASP_Opportunity_TOT/sommapesi))

    app=System.objects.get(appId=appId)

    ScoreAlreadyCreated=ThreatAgentRiskScores.objects.filter(app=app)
    if(not ThreatAgentRiskScores.objects.filter(app=app).exists()):
        obj=ThreatAgentRiskScores.objects.get_or_create(
        app=app,
        skill=OWASP_Skill_TOT,
        size = OWASP_Size_TOT,
        motive = OWASP_Motive_TOT,
        opportunity = OWASP_Opportunity_TOT)

    return render(request, 'stride_impact_evaluation.html', {"appId": appId})
